"""Flatten iw_buildings.txt ↔ flat rows (Excel / JSON)."""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable

from vic3_assign import VIC3_ASSIGN as A, prepare_game_content, read_game_content

BUILDING_KEY_RE = re.compile(rf'building\s*{A}\s*(?:"([^"]+)"|([\w]+))')

BUILDING_SUFFIX_ALIASES = {
    "building_financial_district": "financial_district",
    "building_manor_house": "manor_house",
}


@dataclass
class FlatBuilding:
    country: str
    state: str
    name: str
    level: int
    pm: list[str]
    ownership: str
    owner_tag: str
    owner_state: str = ""
    id: int = 0

    def as_tuple(self) -> tuple[Any, ...]:
        row = normalize_row(self)
        return (
            row.country,
            row.state,
            row.name,
            row.level,
            tuple(row.pm),
            row.ownership,
            row.owner_tag,
            row.owner_state,
        )


def effective_owner_tag(row: FlatBuilding) -> str:
    return row.owner_tag or row.country


def effective_owner_state(row: FlatBuilding) -> str:
    return row.owner_state or row.state


def normalize_row(row: FlatBuilding) -> FlatBuilding:
    """Omit owner fields when they match the building's country / state."""
    if row.owner_state == row.state:
        row.owner_state = ""
    if row.owner_tag == row.country:
        row.owner_tag = ""
    return row


def building_owner_key(row: FlatBuilding) -> tuple[str, str, str, str, str, str]:
    """Identity for one ownership slice of a building in a state."""
    return (
        row.country,
        row.state,
        row.name,
        row.ownership,
        row.owner_tag,
        row.owner_state,
    )


def pm_equal_ignoring_order(left: list[str], right: list[str]) -> bool:
    """Compare PM lists as multisets; order does not matter."""
    if len(left) != len(right):
        return False
    return sorted(left) == sorted(right)


def merge_building_rows(
    rows: list[FlatBuilding],
    *,
    on_warn: Callable[[str], None] | None = None,
) -> list[FlatBuilding]:
    """Merge duplicate rows; ownership may split, PM must match per (tag, state, building)."""

    def warn(message: str) -> None:
        if on_warn is not None:
            on_warn(message)

    canonical_pm: dict[tuple[str, str, str], list[str]] = {}
    grouped: dict[tuple[str, str, str, str, str, str], FlatBuilding] = {}

    for row in rows:
        normalized = FlatBuilding(
            country=row.country,
            state=row.state,
            name=row.name,
            level=row.level,
            pm=list(row.pm),
            ownership=row.ownership,
            owner_tag=row.owner_tag,
            owner_state=row.owner_state,
            id=row.id,
        )
        normalize_row(normalized)
        site_key = (normalized.country, normalized.state, normalized.name)
        owner_key = building_owner_key(normalized)

        pm = list(normalized.pm)
        expected_pm = canonical_pm.get(site_key)
        if expected_pm is None:
            canonical_pm[site_key] = pm
        elif not pm_equal_ignoring_order(expected_pm, pm):
            warn(
                f"building {site_key}：PM 不一致"
                f"（已有 {expected_pm}，当前 {pm}）"
            )
            continue

        existing = grouped.get(owner_key)
        if existing is None:
            grouped[owner_key] = normalized
        else:
            existing.level += normalized.level

    return list(grouped.values())


def _find_block_end(text: str, start: int) -> int:
    from vic3_assign import find_block_end

    return find_block_end(text, start)


def _parse_pm(block: str) -> list[str]:
    m = re.search(rf"activate_production_methods\s*{A}\s*\{{", block)
    if not m:
        return []
    start = m.end() - 1
    end = _find_block_end(block, start)
    inner = block[start + 1:end]
    return re.findall(r"(pm_[\w]+)", inner)


def _extract_country_tag(inner: str) -> str | None:
    m = re.search(rf'country\s*{A}\s*"?c:(\w+)"?', inner)
    return m.group(1) if m else None


def _extract_levels(inner: str) -> int | None:
    m = re.search(rf"levels\s*{A}\s*(\d+)", inner)
    return int(m.group(1)) if m else None


def _extract_type(inner: str) -> str | None:
    m = re.search(rf'type\s*{A}\s*"([^"]+)"', inner)
    if m:
        return m.group(1)
    m = re.search(rf"type\s*{A}\s*([\w]+)", inner)
    return m.group(1) if m else None


def _extract_region(inner: str) -> str | None:
    m = re.search(rf'region\s*{A}\s*"?(STATE_\w+)"?', inner)
    return m.group(1) if m else None


def _ownership_from_building_type(
    owner_type: str, building_name: str, owner_state: str, state: str
) -> str:
    if owner_type == building_name and owner_state == state:
        return "self"
    if owner_type in BUILDING_SUFFIX_ALIASES:
        return BUILDING_SUFFIX_ALIASES[owner_type]
    if owner_type.startswith("building_"):
        return owner_type[len("building_") :]
    return owner_type


def _parse_owner(
    add_ownership: str, building_name: str, state: str
) -> tuple[str, str, str, int]:
    parts: list[tuple[str, str, str, str, int]] = []

    for _key, inner in _iter_assigned_blocks(add_ownership, r"(country)"):
        owner_tag = _extract_country_tag(inner)
        level = _extract_levels(inner)
        if owner_tag is not None and level is not None:
            parts.append(("country", "country", owner_tag, "", level))

    for _key, inner in _iter_assigned_blocks(add_ownership, r"(building)"):
        owner_type = _extract_type(inner)
        owner_tag = _extract_country_tag(inner)
        owner_state = _extract_region(inner) or ""
        level = _extract_levels(inner)
        if owner_type and owner_tag is not None and level is not None:
            ownership = _ownership_from_building_type(
                owner_type, building_name, owner_state, state
            )
            parts.append((ownership, owner_type, owner_tag, owner_state, level))

    for _key, inner in _iter_assigned_blocks(add_ownership, r"(company)"):
        company_type = _extract_type(inner)
        owner_tag = _extract_country_tag(inner)
        level = _extract_levels(inner)
        if company_type and owner_tag is not None and level is not None:
            parts.append((company_type, company_type, owner_tag, "", level))

    if not parts:
        raise ValueError(f"add_ownership 中无法解析所有者：{add_ownership[:120]}")

    total_level = sum(part[4] for part in parts)
    if len(parts) == 1:
        ownership, _owner_type, owner_tag, owner_state, _level = parts[0]
        return ownership, owner_tag, owner_state, total_level

    non_country = [part for part in parts if part[0] != "country"]
    non_self = [part for part in non_country if part[0] != "self"]
    chosen = non_self[0] if non_self else (non_country[0] if non_country else parts[0])
    return chosen[0], chosen[2], chosen[3], total_level


def encode_owner_type(ownership: str, building_name: str) -> str:
    if ownership == "self":
        return building_name
    if ownership == "country":
        raise ValueError("country 所有权无法映射为建筑类型")
    if ownership.startswith("company_"):
        raise ValueError("company 所有权无法映射为建筑类型")
    if ownership in BUILDING_SUFFIX_ALIASES.values():
        for full, short in BUILDING_SUFFIX_ALIASES.items():
            if short == ownership:
                return full
    if ownership.startswith("building_"):
        return ownership
    return f"building_{ownership}"


def _iter_assigned_blocks(inner: str, key_pattern: str) -> list[tuple[str, str]]:
    """Yield (key, block_inner) for each `key = { ... }` at the top level of inner."""
    blocks: list[tuple[str, str]] = []
    i, n = 0, len(inner)
    while i < n:
        m = re.match(r"\s*" + key_pattern + rf"\s*{A}\s*\{{", inner[i:])
        if not m:
            i += 1
            continue
        key = m.group(1)
        start = i + m.end() - 1
        end = _find_block_end(inner, start)
        blocks.append((key, inner[start + 1:end]))
        i = end + 1
    return blocks


def _building_name(block: str) -> str:
    name_m = BUILDING_KEY_RE.search(block)
    if not name_m:
        raise ValueError("create_building 缺少 building 字段")
    return name_m.group(1) or name_m.group(2)


def _extract_reserves(block: str) -> int | None:
    m = re.search(rf"reserves\s*{A}\s*(\d+)", block)
    return int(m.group(1)) if m else None


def _parse_create_building_block(
    block: str, state: str, country: str
) -> FlatBuilding | None:
    reserves = _extract_reserves(block)
    if reserves is not None and reserves != 1:
        return None

    name = _building_name(block)

    ao_m = re.search(rf"add_ownership\s*{A}\s*\{{", block)
    if not ao_m:
        level_m = re.search(rf"level\s*{A}\s*(\d+)", block)
        if not level_m:
            raise ValueError(f"{name} 缺少 add_ownership，且未提供 level")
        return normalize_row(
            FlatBuilding(
                country=country,
                state=state,
                name=name,
                level=int(level_m.group(1)),
                pm=_parse_pm(block),
                ownership="country",
                owner_tag="",
                owner_state="",
            )
        )
    ao_start = ao_m.end() - 1
    ao_end = _find_block_end(block, ao_start)
    add_ownership = block[ao_start + 1:ao_end]

    ownership, owner_tag, owner_state, level = _parse_owner(add_ownership, name, state)

    return normalize_row(
        FlatBuilding(
            country=country,
            state=state,
            name=name,
            level=level,
            pm=_parse_pm(block),
            ownership=ownership,
            owner_tag=owner_tag,
            owner_state=owner_state,
        )
    )


def parse_iw_buildings(text: str) -> list[FlatBuilding]:
    return _parse_buildings_text(text)


def _parse_buildings_text(text: str) -> list[FlatBuilding]:
    text = prepare_game_content(text)
    root_m = re.search(rf"BUILDINGS\s*{A}\s*\{{", text)
    if not root_m:
        return []
    root_start = root_m.end() - 1
    root_end = _find_block_end(text, root_start)
    buildings_inner = text[root_start + 1:root_end]

    rows: list[FlatBuilding] = []
    for state_key, state_inner in _iter_assigned_blocks(
        buildings_inner, r"(s:(STATE_\w+))"
    ):
        state = state_key.split(":")[1]
        for rs_key, rs_inner in _iter_assigned_blocks(
            state_inner, r"(region_state:(\w+))"
        ):
            country = rs_key.split(":")[1]
            pos = 0
            while True:
                cm = re.search(rf"create_building\s*{A}\s*\{{", rs_inner[pos:])
                if not cm:
                    break
                rel = pos + cm.start()
                start = rel + cm.group().index("{")
                end = _find_block_end(rs_inner, start)
                block = rs_inner[start + 1:end]
                row = _parse_create_building_block(block, state, country)
                if row is not None:
                    rows.append(row)
                pos = end + 1

    return rows


def parse_history_buildings_text(text: str) -> list[FlatBuilding]:
    return _parse_buildings_text(text)


def parse_history_buildings_dir(
    buildings_dir: Path | None = None,
    *,
    paths: list[Path] | tuple[Path, ...] | None = None,
) -> list[FlatBuilding]:
    from game_content_resolver import is_empty_content_file, list_txt_files

    if paths is None:
        if buildings_dir is None:
            raise ValueError("必须提供 buildings_dir 或 paths 参数")
        txt_paths = list_txt_files(buildings_dir)
    else:
        txt_paths = list(paths)

    rows: list[FlatBuilding] = []
    for path in txt_paths:
        if is_empty_content_file(path):
            continue
        rows.extend(parse_history_buildings_text(read_game_content(path)))
    return rows


def render_create_building(row: FlatBuilding, indent: str = "\t\t\t") -> str:
    row = normalize_row(row)
    pm_lines = " ".join(row.pm)
    lines = [
        f"{indent}create_building = {{",
        f"{indent}\tbuilding = {row.name}",
        f"{indent}\tadd_ownership = {{",
    ]

    owner_tag = effective_owner_tag(row)

    if row.ownership == "country":
        lines.extend(
            [
                f"{indent}\t\tcountry = {{",
                f"{indent}\t\t\tcountry = c:{owner_tag}",
                f"{indent}\t\t\tlevels = {row.level}",
                f"{indent}\t\t}}",
            ]
        )
    elif row.ownership.startswith("company_"):
        lines.extend(
            [
                f"{indent}\t\tcompany = {{",
                f"{indent}\t\t\ttype = {row.ownership}",
                f"{indent}\t\t\tcountry = c:{owner_tag}",
                f"{indent}\t\t\tlevels = {row.level}",
                f"{indent}\t\t}}",
            ]
        )
    else:
        owner_type = encode_owner_type(row.ownership, row.name)
        owner_state = effective_owner_state(row)
        lines.extend(
            [
                f"{indent}\t\tbuilding = {{",
                f"{indent}\t\t\ttype = {owner_type}",
                f"{indent}\t\t\tcountry = c:{owner_tag}",
                f"{indent}\t\t\tlevels = {row.level}",
                f"{indent}\t\t\tregion = {owner_state}",
                f"{indent}\t\t}}",
            ]
        )

    lines.extend(
        [
            f"{indent}\t}}",
            f"{indent}\treserves = 1",
            f"{indent}\tactivate_production_methods = {{",
            f"{indent}\t\t{pm_lines} ",
            f"{indent}\t}}",
            f"{indent}}}",
        ]
    )
    return "\n".join(lines)


def render_iw_buildings(rows: list[FlatBuilding]) -> str:
    """Regenerate BUILDINGS block preserving row order within each region_state."""
    state_order: list[str] = []
    country_order: dict[str, list[str]] = {}
    groups: dict[str, dict[str, list[FlatBuilding]]] = {}

    for row in rows:
        if row.state not in groups:
            groups[row.state] = {}
            state_order.append(row.state)
            country_order[row.state] = []
        if row.country not in groups[row.state]:
            groups[row.state][row.country] = []
            country_order[row.state].append(row.country)
        groups[row.state][row.country].append(row)

    lines = ["BUILDINGS = {"]
    for state in state_order:
        lines.append(f"\ts:{state} = {{")
        for country in country_order[state]:
            lines.append(f"\t\tregion_state:{country} = {{")
            for row in groups[state][country]:
                lines.append(render_create_building(row, indent="\t\t\t"))
            lines.append("\t\t}")
        lines.append("\t}")
    lines.append("}")
    return "\n".join(lines) + "\n"


def assign_ids(rows: list[FlatBuilding], start: int = 1) -> list[FlatBuilding]:
    for i, row in enumerate(rows, start=start):
        row.id = i
    return rows


def rows_to_json(rows: list[FlatBuilding]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for r in rows:
        row = normalize_row(r)
        out.append(
            {
                "id": row.id,
                "country": row.country,
                "state": row.state,
                "name": row.name,
                "level": row.level,
                "pm": row.pm,
                "ownership": row.ownership,
                "owner_tag": row.owner_tag,
                "owner_state": row.owner_state,
            }
        )
    return out


def load_excel(path: Path, *, sheet: str = "buildings") -> list[FlatBuilding]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet in wb.sheetnames else wb.active
        data = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not data:
        raise ValueError(f"工作表为空：{path}")

    header = [str(col) for col in data[0]]
    rows: list[FlatBuilding] = []
    for line in data[1:]:
        record = dict(zip(header, line))
        pm_raw = record.get("pm")
        if isinstance(pm_raw, str) and pm_raw.strip():
            pm = json.loads(pm_raw)
        elif isinstance(pm_raw, list):
            pm = [str(item) for item in pm_raw]
        else:
            pm = []
        rows.append(
            normalize_row(
                FlatBuilding(
                    country=str(record["country"]),
                    state=str(record["state"]),
                    name=str(record["name"]),
                    level=int(record["level"]),
                    pm=[str(p) for p in pm],
                    ownership=str(record["ownership"]),
                    owner_tag=str(record.get("owner_tag") or ""),
                    owner_state=str(record.get("owner_state") or ""),
                    id=int(record.get("id") or 0),
                )
            )
        )
    return rows


def export_excel(rows: list[FlatBuilding], path: Path) -> Path:
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "buildings"
    headers = [
        "id",
        "country",
        "state",
        "name",
        "level",
        "pm",
        "ownership",
        "owner_tag",
        "owner_state",
    ]
    ws.append(headers)
    for r in rows:
        row = normalize_row(r)
        ws.append(
            [
                row.id,
                row.country,
                row.state,
                row.name,
                row.level,
                json.dumps(row.pm, ensure_ascii=False),
                row.ownership,
                row.owner_tag,
                row.owner_state,
            ]
        )
    wb.save(path)
    return path


def roundtrip_check(source_path: Path) -> dict[str, Any]:
    text = source_path.read_text(encoding="utf-8")
    rows = parse_iw_buildings(text)
    assign_ids(rows)
    rendered = render_iw_buildings(rows)
    rows2 = parse_iw_buildings(rendered)

    orig_tuples = [r.as_tuple() for r in rows]
    new_tuples = [r.as_tuple() for r in rows2]

    return {
        "source_rows": len(rows),
        "roundtrip_rows": len(rows2),
        "content_match": orig_tuples == new_tuples,
        "rows": rows,
        "rendered": rendered,
    }

