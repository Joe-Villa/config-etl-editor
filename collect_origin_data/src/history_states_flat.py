"""Serialize / deserialize common/history/states ↔ flat tables (lossless logic)."""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import TYPE_CHECKING, Any, Iterator

if TYPE_CHECKING:
    from history_pops_flat import PopByTagRow

from vic3_assign import VIC3_ASSIGN as A, block_header, prepare_game_content, read_game_content

STATE_KEY_RE = re.compile(
    block_header(r"STATE_\w+", line_prefix="s:"),
    flags=re.MULTILINE,
)
STATE_BLOCK_HEADER_RE = re.compile(
    block_header(r"STATE_\w+", line_prefix="s:"),
    flags=re.MULTILINE,
)
CREATE_STATE_RE = re.compile(rf"create_state\s*{A}\s*\{{")
ADD_HOMELAND_RE = re.compile(rf"add_homeland\s*{A}\s*(cu:\w+)")
ADD_CLAIM_RE = re.compile(rf"add_claim\s*{A}\s*(c:\w+)")
COUNTRY_RE = re.compile(rf"country\s*{A}\s*(c:\w+)", re.IGNORECASE)
STATE_TYPE_RE = re.compile(rf"state_type\s*{A}\s*(\w+)")
OWNED_PROVINCES_RE = re.compile(rf"owned_provinces\s*{A}\s*\{{")
_KEY_ASSIGN_RE = re.compile(rf"([A-Za-z_][A-Za-z0-9_]*)\s*{A}")

STATE_BLOCK_STANDARD_KEYS = frozenset({"create_state", "add_homeland", "add_claim"})
CREATE_STATE_STANDARD_KEYS = frozenset({"country", "owned_provinces", "state_type"})


@dataclass
class StateMetaRow:
    state: str
    homelands: list[str] = field(default_factory=list)
    claims: list[str] = field(default_factory=list)


@dataclass
class StateOwnershipRow:
    state: str
    tag: str
    owned_provinces: list[str] = field(default_factory=list)
    state_type: str = "incorporated"
    population: int = 0


def _find_block_end(text: str, start: int) -> int:
    from vic3_assign import find_block_end

    return find_block_end(text, start)


def _strip_prefix(value: str, prefix: str) -> str:
    return value[len(prefix) :] if value.startswith(prefix) else value


def _split_csv(value: str) -> list[str]:
    if not value or not str(value).strip():
        return []
    return [part.strip() for part in str(value).split(",") if part.strip()]


def _join_csv(items: list[str]) -> str:
    return ",".join(items)


def _parse_owned_provinces(block: str) -> list[str]:
    m = OWNED_PROVINCES_RE.search(block)
    if not m:
        return []
    start = m.end() - 1
    end = _find_block_end(block, start)
    inner = block[start + 1 : end]
    return re.findall(r"(x[\w]+)", inner, flags=re.IGNORECASE)


def _parse_create_state(block: str, state: str) -> StateOwnershipRow | None:
    country_m = COUNTRY_RE.search(block)
    if not country_m:
        return None
    tag = country_m.group(1)[2:]
    state_type = "incorporated"
    type_m = STATE_TYPE_RE.search(block)
    if type_m:
        state_type = type_m.group(1)
    return StateOwnershipRow(
        state=state,
        tag=tag,
        owned_provinces=_parse_owned_provinces(block),
        state_type=state_type,
    )


def _parse_state_block(state: str, block: str) -> tuple[StateMetaRow | None, list[StateOwnershipRow]]:
    if not CREATE_STATE_RE.search(block):
        return None, []
    homelands = [_strip_prefix(v, "cu:") for v in ADD_HOMELAND_RE.findall(block)]
    claims = [_strip_prefix(v, "c:") for v in ADD_CLAIM_RE.findall(block)]
    ownership: list[StateOwnershipRow] = []
    for m in CREATE_STATE_RE.finditer(block):
        cs_start = m.end() - 1
        cs_end = _find_block_end(block, cs_start)
        row = _parse_create_state(block[cs_start + 1 : cs_end], state)
        if row is not None:
            ownership.append(row)
    return StateMetaRow(state=state, homelands=homelands, claims=claims), ownership


def _state_block_inner_has_create_state(block_inner: str) -> bool:
    return CREATE_STATE_RE.search(block_inner) is not None


def _inject_into_state_block(full_block: str, additive_inner: str) -> str:
    """Append additive inner script before the closing ``}`` of a state block."""
    extra = additive_inner.strip()
    if not extra:
        return full_block
    close_at = full_block.rfind("}")
    if close_at < 0:
        return full_block
    return full_block[:close_at] + "\n" + extra + "\n" + full_block[close_at:]


def merge_history_states_blocks(
    paths: list[Path] | tuple[Path, ...],
    mod_dir: Path,
    *,
    skip_empty: bool = True,
) -> dict[str, str]:
    """Merge history state blocks; additive files without ``create_state`` inject into prior block."""
    from game_content_resolver import (
        is_empty_content_file,
        iter_paradox_blocks,
        ordered_merge_paths,
    )

    blocks: dict[str, str] = {}
    for path in ordered_merge_paths(paths, mod_dir):
        if skip_empty and is_empty_content_file(path):
            continue
        text = read_game_content(path)
        for key, full_block in iter_paradox_blocks(
            text,
            r"STATE_\w+",
            line_prefix="s:",
        ):
            inner_start = full_block.index("{") + 1
            inner_end = full_block.rfind("}")
            inner = full_block[inner_start:inner_end]
            if _state_block_inner_has_create_state(inner):
                blocks[key] = full_block
            elif key in blocks:
                blocks[key] = _inject_into_state_block(blocks[key], inner)
            else:
                blocks[key] = full_block
    return blocks


def read_merged_history_states_blocks(
    paths: list[Path] | tuple[Path, ...],
    mod_dir: Path,
    *,
    skip_empty: bool = True,
) -> str:
    blocks = merge_history_states_blocks(
        paths,
        mod_dir,
        skip_empty=skip_empty,
    )
    if not blocks:
        return ""
    return "\n".join(blocks.values()) + "\n"


def parse_states_text(text: str) -> tuple[list[StateMetaRow], list[StateOwnershipRow]]:
    text = prepare_game_content(text)
    meta_rows: list[StateMetaRow] = []
    ownership_rows: list[StateOwnershipRow] = []
    for m in STATE_KEY_RE.finditer(text):
        state = m.group(1)
        block_start = m.end() - 1
        block_end = _find_block_end(text, block_start)
        block = text[block_start + 1 : block_end]
        meta, own = _parse_state_block(state, block)
        if meta is None:
            continue
        meta_rows.append(meta)
        ownership_rows.extend(own)
    return meta_rows, ownership_rows


def parse_states_file(path: Path) -> tuple[list[StateMetaRow], list[StateOwnershipRow]]:
    return parse_states_text(read_game_content(path))


_NO_CREATE_STATE_SKIP_REASON = (
    "块内无 create_state，本程序不处理此类写法，已跳过"
)


@dataclass(frozen=True)
class _DirectAssignment:
    key: str
    key_start: int
    value_brace: int | None
    end: int


def _iter_direct_assignments(block_inner: str) -> Iterator[_DirectAssignment]:
    """Yield direct ``key = ...`` children of a Paradox block body."""
    i = 0
    n = len(block_inner)
    while i < n:
        while i < n and block_inner[i].isspace():
            i += 1
        if i >= n:
            break
        match = _KEY_ASSIGN_RE.match(block_inner, i)
        if not match:
            i += 1
            continue
        key = match.group(1)
        key_start = i
        i = match.end()
        while i < n and block_inner[i].isspace():
            i += 1
        if i < n and block_inner[i] == "{":
            value_brace = i
            end = _find_block_end(block_inner, value_brace) + 1
            i = end
            yield _DirectAssignment(key, key_start, value_brace, end)
            continue
        line_end = block_inner.find("\n", i)
        end = n if line_end < 0 else line_end + 1
        i = end
        yield _DirectAssignment(key, key_start, None, end)


def _non_standard_stmt_reason(key: str) -> str:
    return f"含非标准语句 {key}，本程序不处理此类写法，已忽略"


def _scan_block_non_standard_assignments(
    block_inner: str,
    block_inner_offset: int,
    state: str,
    *,
    context_label: str,
    standard_keys: frozenset[str],
    text: str,
    source: str,
    relative_dir: str,
    filename: str,
    log: object,
) -> None:
    from import_context import format_import_warning, line_at

    for assignment in _iter_direct_assignments(block_inner):
        if assignment.key in standard_keys:
            if assignment.key == "create_state" and assignment.value_brace is not None:
                cs_inner = block_inner[
                    assignment.value_brace + 1 : _find_block_end(block_inner, assignment.value_brace)
                ]
                _scan_block_non_standard_assignments(
                    cs_inner,
                    block_inner_offset + assignment.value_brace + 1,
                    state,
                    context_label="create_state",
                    standard_keys=CREATE_STATE_STANDARD_KEYS,
                    text=text,
                    source=source,
                    relative_dir=relative_dir,
                    filename=filename,
                    log=log,
                )
            continue
        line = line_at(text, block_inner_offset + assignment.key_start)
        log.warn(
            format_import_warning(
                source,
                relative_dir,
                filename,
                line,
                f"解析开局{context_label}（州 {state}）",
                _non_standard_stmt_reason(assignment.key),
            )
        )


def scan_states_file_no_create_state_warnings(
    path: Path,
    mod_root: Path,
    vanilla: Path,
    log: object,
) -> None:
    """Warn when a state block has no ``create_state`` and will be skipped on import."""
    from import_context import classify_content_path, format_import_warning, line_at

    text = read_game_content(path)
    source, relative_dir, filename = classify_content_path(path, mod_root, vanilla)
    for match in STATE_BLOCK_HEADER_RE.finditer(text):
        state = match.group(1)
        block_start = match.end() - 1
        block_end = _find_block_end(text, block_start)
        block = text[block_start + 1 : block_end]
        if _state_block_inner_has_create_state(block):
            continue
        line = line_at(text, match.start())
        log.warn(
            format_import_warning(
                source,
                relative_dir,
                filename,
                line,
                f"解析开局州历史块（州 {state}）",
                _NO_CREATE_STATE_SKIP_REASON,
            )
        )


def scan_states_paths_no_create_state_warnings(
    paths: list[Path] | tuple[Path, ...],
    mod_dir: Path,
    mod_root: Path,
    vanilla: Path,
    log: object,
) -> None:
    """Warn for standalone state blocks without ``create_state`` after merge simulation."""
    from game_content_resolver import is_empty_content_file, ordered_merge_paths
    from import_context import classify_content_path, format_import_warning, line_at

    blocks: dict[str, str] = {}
    for path in ordered_merge_paths(paths, mod_dir):
        if is_empty_content_file(path):
            continue
        text = read_game_content(path)
        source, relative_dir, filename = classify_content_path(path, mod_root, vanilla)
        for match in STATE_BLOCK_HEADER_RE.finditer(text):
            key = match.group(1)
            block_start = match.end() - 1
            block_end = _find_block_end(text, block_start)
            full_block = text[match.start() : block_end + 1]
            inner_start = full_block.index("{") + 1
            inner_end = full_block.rfind("}")
            inner = full_block[inner_start:inner_end]
            if _state_block_inner_has_create_state(inner):
                blocks[key] = full_block
            elif key in blocks:
                blocks[key] = _inject_into_state_block(blocks[key], inner)
            else:
                blocks[key] = full_block
                line = line_at(text, match.start())
                log.warn(
                    format_import_warning(
                        source,
                        relative_dir,
                        filename,
                        line,
                        f"解析开局州历史块（州 {key}）",
                        _NO_CREATE_STATE_SKIP_REASON,
                    )
                )


def scan_states_file_inner_non_standard_warnings(
    path: Path,
    mod_root: Path,
    vanilla: Path,
    log: object,
) -> None:
    """Warn for non-standard statements inside state / create_state blocks."""
    from import_context import classify_content_path

    text = read_game_content(path)
    source, relative_dir, filename = classify_content_path(path, mod_root, vanilla)
    for match in STATE_BLOCK_HEADER_RE.finditer(text):
        state = match.group(1)
        block_start = match.end() - 1
        block_end = _find_block_end(text, block_start)
        block_inner = text[block_start + 1 : block_end]
        _scan_block_non_standard_assignments(
            block_inner,
            block_start + 1,
            state,
            context_label="州历史块",
            standard_keys=STATE_BLOCK_STANDARD_KEYS,
            text=text,
            source=source,
            relative_dir=relative_dir,
            filename=filename,
            log=log,
        )


def scan_states_paths_inner_non_standard_warnings(
    paths: list[Path] | tuple[Path, ...],
    mod_dir: Path,
    mod_root: Path,
    vanilla: Path,
    log: object,
) -> None:
    from game_content_resolver import is_empty_content_file, ordered_merge_paths

    for path in ordered_merge_paths(paths, mod_dir):
        if is_empty_content_file(path):
            continue
        scan_states_file_inner_non_standard_warnings(path, mod_root, vanilla, log)


def scan_states_file_create_state_errors(
    path: Path,
    mod_root: Path,
    vanilla: Path,
    log: object,
) -> None:
    """Log import errors for create_state blocks missing country (with file line)."""
    from import_context import classify_content_path, format_import_error, line_at

    text = read_game_content(path)
    source, relative_dir, filename = classify_content_path(path, mod_root, vanilla)
    for sm in STATE_KEY_RE.finditer(text):
        state = sm.group(1)
        block_start = sm.end() - 1
        block_end = _find_block_end(text, block_start)
        block = text[block_start + 1 : block_end]
        block_offset = block_start + 1
        for cm in CREATE_STATE_RE.finditer(block):
            cs_start = cm.end() - 1
            cs_end = _find_block_end(block, cs_start)
            inner = block[cs_start + 1 : cs_end]
            if COUNTRY_RE.search(inner):
                continue
            abs_index = sm.start() + block_offset + cm.start()
            line = line_at(text, abs_index)
            log.error(
                format_import_error(
                    source,
                    relative_dir,
                    filename,
                    line,
                    f"解析开局州历史 create_state（州 {state}）",
                    "缺少 country 字段，无法确定归属 tag",
                )
            )


def scan_states_file_brace_errors(
    path: Path,
    mod_root: Path,
    vanilla: Path,
    log: object,
) -> None:
    """Log import errors when a state block swallows following states (brace mismatch)."""
    from import_context import classify_content_path, format_import_error, line_at

    text = read_game_content(path)
    source, relative_dir, filename = classify_content_path(path, mod_root, vanilla)
    inner_header_re = re.compile(
        block_header(r"STATE_\w+", line_prefix="s:"),
        flags=re.MULTILINE,
    )
    for match in STATE_BLOCK_HEADER_RE.finditer(text):
        state = match.group(1)
        block_start = match.end() - 1
        try:
            block_end = _find_block_end(text, block_start)
        except ValueError:
            line = line_at(text, match.start())
            log.error(
                format_import_error(
                    source,
                    relative_dir,
                    filename,
                    line,
                    f"解析开局州历史块（州 {state}）",
                    "括号不匹配，存在未闭合的 { } 块",
                )
            )
            continue
        inner = text[block_start + 1 : block_end]
        swallowed = [m.group(1) for m in inner_header_re.finditer(inner)]
        if not swallowed:
            continue
        line = line_at(text, match.start())
        preview = ", ".join(swallowed[:5])
        if len(swallowed) > 5:
            preview = f"{preview} 等 {len(swallowed)} 个"
        log.error(
            format_import_error(
                source,
                relative_dir,
                filename,
                line,
                f"解析开局州历史块（州 {state}）",
                f"括号不匹配，块内包含本应独立的州定义：{preview}",
            )
        )


def scan_states_paths_brace_errors(
    paths: list[Path] | tuple[Path, ...],
    mod_dir: Path,
    mod_root: Path,
    vanilla: Path,
    log: object,
) -> None:
    from game_content_resolver import is_empty_content_file, ordered_merge_paths

    for path in ordered_merge_paths(paths, mod_dir):
        if is_empty_content_file(path):
            continue
        scan_states_file_brace_errors(path, mod_root, vanilla, log)


def scan_states_paths_create_state_errors(
    paths: list[Path] | tuple[Path, ...],
    mod_dir: Path,
    mod_root: Path,
    vanilla: Path,
    log: object,
) -> None:
    from game_content_resolver import is_empty_content_file, ordered_merge_paths

    for path in ordered_merge_paths(paths, mod_dir):
        if is_empty_content_file(path):
            continue
        scan_states_file_create_state_errors(path, mod_root, vanilla, log)


def parse_states_dir(
    states_dir: Path | None = None,
    *,
    paths: list[Path] | tuple[Path, ...] | None = None,
    mod_dir: Path | None = None,
    mod_root: Path | None = None,
    vanilla: Path | None = None,
    log: object | None = None,
) -> tuple[list[StateMetaRow], list[StateOwnershipRow]]:
    from game_content_resolver import is_empty_content_file, list_txt_files

    if paths is None:
        if states_dir is None:
            raise ValueError("必须提供 states_dir 或 paths 参数")
        txt_paths = list_txt_files(states_dir)
    else:
        txt_paths = list(paths)

    if mod_dir is not None:
        if log is not None and mod_root is not None and vanilla is not None:
            scan_states_paths_create_state_errors(
                txt_paths, mod_dir, mod_root, vanilla, log
            )
            scan_states_paths_brace_errors(
                txt_paths, mod_dir, mod_root, vanilla, log
            )
            scan_states_paths_no_create_state_warnings(
                txt_paths, mod_dir, mod_root, vanilla, log
            )
            scan_states_paths_inner_non_standard_warnings(
                txt_paths, mod_dir, mod_root, vanilla, log
            )
        text = read_merged_history_states_blocks(
            txt_paths,
            mod_dir,
        )
        meta_rows, ownership_rows = parse_states_text(text)
    else:
        meta_by: dict[str, StateMetaRow] = {}
        own_by: dict[tuple[str, str], StateOwnershipRow] = {}
        for path in txt_paths:
            if is_empty_content_file(path):
                continue
            if log is not None and mod_root is not None and vanilla is not None:
                scan_states_file_create_state_errors(path, mod_root, vanilla, log)
                scan_states_file_brace_errors(path, mod_root, vanilla, log)
                scan_states_file_no_create_state_warnings(path, mod_root, vanilla, log)
                scan_states_file_inner_non_standard_warnings(path, mod_root, vanilla, log)
            meta, own = parse_states_file(path)
            for row in meta:
                meta_by[row.state] = row
            for row in own:
                own_by[(row.state, row.tag)] = row
        meta_rows = sorted(meta_by.values(), key=lambda r: r.state)
        ownership_rows = sorted(own_by.values(), key=lambda r: (r.state, r.tag))
        return meta_rows, ownership_rows

    return meta_rows, ownership_rows


def _state_order(meta_rows: list[StateMetaRow], ownership_rows: list[StateOwnershipRow]) -> list[str]:
    order: list[str] = []
    seen: set[str] = set()
    for row in meta_rows:
        if row.state not in seen:
            order.append(row.state)
            seen.add(row.state)
    for row in ownership_rows:
        if row.state not in seen:
            order.append(row.state)
            seen.add(row.state)
    return order


def render_states(
    meta_rows: list[StateMetaRow],
    ownership_rows: list[StateOwnershipRow],
) -> str:
    meta_by = {row.state: row for row in meta_rows}
    own_by: dict[str, list[StateOwnershipRow]] = {}
    for row in ownership_rows:
        own_by.setdefault(row.state, []).append(row)

    lines = ["STATES = {"]
    for state in _state_order(meta_rows, ownership_rows):
        lines.append(f"\ts:{state} = {{")
        for own in own_by.get(state, []):
            lines.append("\t\tcreate_state = {")
            lines.append(f"\t\t\tcountry = c:{own.tag}")
            if own.state_type == "unincorporated":
                lines.append("\t\t\tstate_type = unincorporated")
            provinces = " ".join(own.owned_provinces)
            lines.append(f"\t\t\towned_provinces = {{ {provinces} }}")
            lines.append("\t\t}")
            lines.append("")
        meta = meta_by.get(state, StateMetaRow(state=state))
        for homeland in meta.homelands:
            lines.append(f"\t\tadd_homeland = cu:{homeland}")
        for claim in meta.claims:
            lines.append(f"\t\tadd_claim = c:{claim}")
        lines.append("\t}")
    lines.append("}")
    return "\n".join(lines) + "\n"


def _meta_to_dict(row: StateMetaRow) -> dict[str, Any]:
    return {
        "state": row.state,
        "homelands": _join_csv(row.homelands),
        "claims": _join_csv(row.claims),
    }


def _ownership_to_dict(row: StateOwnershipRow) -> dict[str, Any]:
    return {
        "state": row.state,
        "tag": row.tag,
        "owned_provinces": _join_csv(row.owned_provinces),
        "state_type": row.state_type,
        "population": row.population,
    }


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() == "none" else text


def meta_from_dict(data: dict[str, Any]) -> StateMetaRow:
    return StateMetaRow(
        state=_cell_text(data["state"]),
        homelands=_split_csv(_cell_text(data.get("homelands", ""))),
        claims=_split_csv(_cell_text(data.get("claims", ""))),
    )


def ownership_from_dict(data: dict[str, Any]) -> StateOwnershipRow:
    state_type = _cell_text(data.get("state_type", "incorporated")) or "incorporated"
    population_raw = data.get("population", 0)
    population = int(population_raw) if population_raw not in (None, "") else 0
    return StateOwnershipRow(
        state=_cell_text(data["state"]),
        tag=_cell_text(data["tag"]),
        owned_provinces=_split_csv(_cell_text(data.get("owned_provinces", ""))),
        state_type=state_type,
        population=population,
    )


def merge_population_into_ownership(
    ownership_rows: list[StateOwnershipRow],
    pops: list["PopByTagRow"],
) -> list[StateOwnershipRow]:
    """Join pops on (state, tag). Missing pops default to 0; orphan pops are ignored."""
    pop_by_key = {(row.state, row.tag): row.population for row in pops}
    return [
        StateOwnershipRow(
            state=row.state,
            tag=row.tag,
            owned_provinces=row.owned_provinces,
            state_type=row.state_type,
            population=pop_by_key.get((row.state, row.tag), 0),
        )
        for row in ownership_rows
    ]


def export_excel(
    meta_rows: list[StateMetaRow],
    ownership_rows: list[StateOwnershipRow],
    path: Path,
) -> Path:
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "meta"
    ws_meta.append(["state", "homelands", "claims"])
    for row in meta_rows:
        ws_meta.append([row.state, _join_csv(row.homelands), _join_csv(row.claims)])

    ws_own = wb.create_sheet("ownership")
    ws_own.append(["state", "tag", "owned_provinces", "state_type", "population"])
    for row in ownership_rows:
        ws_own.append(
            [
                row.state,
                row.tag,
                _join_csv(row.owned_provinces),
                row.state_type,
                row.population,
            ]
        )
    wb.save(path)
    return path


def load_excel(path: Path) -> tuple[list[StateMetaRow], list[StateOwnershipRow]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    meta_rows: list[StateMetaRow] = []
    own_rows: list[StateOwnershipRow] = []

    ws_meta = wb["meta"]
    meta_headers = [cell.value for cell in next(ws_meta.iter_rows(min_row=1, max_row=1))]
    for row in ws_meta.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        data = dict(zip(meta_headers, row))
        meta_rows.append(meta_from_dict(data))

    ws_own = wb["ownership"]
    own_headers = [cell.value for cell in next(ws_own.iter_rows(min_row=1, max_row=1))]
    for row in ws_own.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        data = dict(zip(own_headers, row))
        own_rows.append(ownership_from_dict(data))
    wb.close()
    return meta_rows, own_rows


def export_json(
    meta_rows: list[StateMetaRow],
    ownership_rows: list[StateOwnershipRow],
    path: Path,
) -> Path:
    payload = {
        "meta": [_meta_to_dict(r) for r in meta_rows],
        "ownership": [_ownership_to_dict(r) for r in ownership_rows],
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return path


def roundtrip_check(source: Path) -> dict[str, Any]:
    meta1, own1 = parse_states_file(source)
    rendered = render_states(meta1, own1)
    meta2, own2 = parse_states_text(rendered)
    return {
        "meta_rows": len(meta1),
        "ownership_rows": len(own1),
        "meta_match": meta1 == meta2,
        "ownership_match": own1 == own2,
        "content_match": meta1 == meta2 and own1 == own2,
        "rendered": rendered,
    }
