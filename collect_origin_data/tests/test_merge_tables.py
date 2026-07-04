from __future__ import annotations

import sqlite3
import sys
import tempfile
import unittest
from pathlib import Path

TOOL_DIR = Path(__file__).resolve().parent.parent / "tool"
SRC_DIR = Path(__file__).resolve().parent.parent / "src"
if str(TOOL_DIR) not in sys.path:
    sys.path.insert(0, str(TOOL_DIR))
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from constants import ORIGIN_DB  # noqa: E402
from merge_tables import export_all  # noqa: E402


def _write_origin_db(path: Path) -> None:
    conn = sqlite3.connect(path)
    conn.execute("CREATE TABLE alpha (a INTEGER)")
    conn.execute("INSERT INTO alpha VALUES (1)")
    conn.commit()
    conn.close()


class MergeTablesTests(unittest.TestCase):
    def test_export_preserves_table_names_as_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            data_dir = root / "data" / "sample"
            data_dir.mkdir(parents=True)
            table_dir = root / "table"

            _write_origin_db(data_dir / ORIGIN_DB)
            (data_dir / "metadata.json").write_text("{}", encoding="utf-8")

            results = export_all(data_dir=root / "data", table_dir=table_dir)
            self.assertEqual(len(results), 1)
            run_id, output_path, db_count, sheet_count = results[0]
            self.assertEqual(run_id, "sample")
            self.assertEqual(db_count, 1)
            self.assertEqual(sheet_count, 1)
            self.assertTrue(output_path.is_file())

            from openpyxl import load_workbook

            merged = load_workbook(output_path, read_only=True)
            try:
                self.assertEqual(merged.sheetnames, ["alpha"])
            finally:
                merged.close()


if __name__ == "__main__":
    unittest.main()
